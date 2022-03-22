# Author : 柠檬班-亚萌
# Project : scb27
# Time : 2022/3/18 20:14
# E-mail : 3343787213@qq.com
# Company : 湖南零檬信息技术有限公司
# Site : http://www.lemonban.com
# Forum : http://testingpai.com

# 接口自动化测试的基本步骤：
# 1、excel测试用例准备好，通过代码读取测试数据  ==  read_data()
# 2、发送接口请求，得到响应结果   == func_req()
# 3、实际结果   vs   预期结果
# 4、把最终的结果回写到excel   == write_data()

import openpyxl
import requests

def read_data(filename,sheet):
    wb = openpyxl.load_workbook(filename)  # 加载excel
    sheet = wb[sheet]  # 获取sheet
    max_row = sheet.max_row  # 获取这个sheet里最大行数
    case_list = []  # 定义一个空列表，用来装读取出来的数据
    for i in range(2,max_row+1,1):  # 取头不取尾
        dict1 = dict(
        cell_id = sheet.cell(row=i,column=1).value,  # 取出用例id
        cell_url = sheet.cell(row=i, column=6).value,  # 取出接口地址
        cell_header = sheet.cell(row=i, column=5).value,  # 取出请求头
        cell_body = sheet.cell(row=i, column=7).value,  # 取出请求体
        cell_expect = sheet.cell(row=i, column=8).value)  # 取出预期结果
        case_list.append(dict1)  # 通过apepend方法，一条一条的用例放里面追加
    return case_list
def func_req(url,body,headers):
    res = requests.post(url=url,json=body,headers=headers)
    res_log = res.json()
    return res_log
def write_data(filename,sheet,row,column,final_result):
    wb = openpyxl.load_workbook(filename)  # 加载excel
    sheet = wb[sheet]  # 获取sheet
    sheet.cell(row=row,column=column).value = final_result  # 写入结果
    wb.save(filename)


def execute_function(filename,sheetname):  #　最后再把代码封装成功函数，以便所有的功能都可以直接调用这个函数就可以完成自动化了
    res = read_data(filename,sheetname)  # 调用读取excel函数，把测试数据读取出来
    for case in res:
        url_login = case.get('cell_url')  # 从读取函数返回的结果里，把需要的接口地址取出来
        header_login = case.get('cell_header')  # 取出请求头
        body_login = case.get('cell_body')  # 取出请求体
        expect_login = case.get('cell_expect')  # 取出预期结果
        id_login = case.get('cell_id')  # 取出用例id
        body_login = eval(body_login)  # 通过eval函数，把读出来的字符串数据转换为原本类型数据
        header_login = eval(header_login) # 通过eval函数，把读出来的字符串数据转换为原本类型数据
        expect_login = eval(expect_login) # 通过eval函数，把读出来的字符串数据转换为原本类型数据
        res = func_req(url=url_login,body=body_login,headers=header_login)  # 调用发送请求函数
        expect_msg = expect_login.get('msg')  # 取出预期结果中的msg
        real_msg = res.get('msg')  # 取出实际结果中的msg
        print('{}功能的执行结果为：{}'.format(sheetname,real_msg))  # 美化一下控制台输出的结果
        if expect_msg == real_msg:  # 做结果判断
            print("第{}条用例通过！！".format(id_login))  # 美化一下控制台输出的结果
            print('*' * 30)  # 美化一下控制台输出的结果
            final_result = 'pass'  # 用变量来标识最终结果
        else:
            print('第{}条用例不通过！！'.format(id_login))  # 美化一下控制台输出的结果
            print('*' * 30)  # 美化一下控制台输出的结果
            final_result = 'fail'  # 用变量来标识最终结果
        write_data(filename,sheetname,id_login+1,9,final_result)  # 调用写入结果函数，把最终结果写回到excel

execute_function('testcase_api_wuye.xlsx','login')
execute_function('testcase_api_wuye.xlsx','register')

